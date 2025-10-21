import openpyxl
from datetime import datetime
import os


def extract_birthday_from_id(id_card):
    """
    从身份证号码中提取出生日期
    """
    if id_card is None or id_card == '':
        return ''

    id_str = str(id_card).strip()

    # 处理15位或18位身份证
    if len(id_str) == 15:
        # 15位身份证：7-12位是出生日期（YYMMDD）
        birthday_str = '19' + id_str[6:12]  # 添加19前缀
    elif len(id_str) == 18:
        # 18位身份证：7-14位是出生日期（YYYYMMDD）
        birthday_str = id_str[6:14]
    else:
        print(f"警告：身份证号码长度不正确: {id_str}")
        return ''

    try:
        # 将字符串转换为日期对象，然后格式化为YYYY-MM-DD
        birthday = datetime.strptime(birthday_str, '%Y%m%d')
        return birthday.strftime('%Y-%m-%d')
    except ValueError as e:
        print(f"警告：无效的出生日期格式: {birthday_str}, 错误: {e}")
        return ''


def process_excel_file():
    """
    处理Excel文件，从身份证号码提取出生日期，保持原有格式
    """
    # 获取文件路径
    file_path = input("请输入Excel文件路径: ").strip().strip('"')

    # 检查文件是否存在
    if not os.path.exists(file_path):
        print("错误：文件不存在！")
        return

    try:
        # 使用openpyxl加载工作簿，保持原有格式
        print("正在加载Excel文件...")
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # 获取最大行数
        max_row = sheet.max_row

        print(f"找到 {max_row} 行数据")

        # 处理每一行的身份证号码
        processed_count = 0
        for row in range(1, max_row + 1):
            # 获取第4列的值（身份证号码）
            id_cell = sheet.cell(row=row, column=4)
            id_card = id_cell.value

            # 提取出生日期
            birthday = extract_birthday_from_id(id_card)

            if birthday:
                # 在第9列写入出生日期
                birthday_cell = sheet.cell(row=row, column=9)
                birthday_cell.value = birthday

                # 如果需要，可以设置日期格式
                # birthday_cell.number_format = 'YYYY-MM-DD'

                processed_count += 1

            # 显示进度
            if row % 100 == 0:
                print(f"已处理 {row} 行...")

        # 生成输出文件名
        base_name = os.path.splitext(file_path)[0]
        output_file = f"{base_name}_with_birthday.xlsx"

        # 保存处理后的文件，保持原有格式
        print("正在保存文件...")
        workbook.save(output_file)

        print(f"\n处理完成！")
        print(f"成功处理: {processed_count} 条记录")
        print(f"输出文件: {output_file}")

    except Exception as e:
        print(f"处理过程中出现错误: {e}")


def main():
    """
    主函数
    """
    print("=" * 50)
    print("   身份证出生日期提取工具")
    print("=" * 50)
    print("功能说明:")
    print("- 从Excel第4列读取身份证号码")
    print("- 在第9列生成格式为YYYY-MM-DD的出生日期")
    print("- 自动处理15位和18位身份证")
    print("- 保持原有Excel文件的格式和样式")
    print("- 生成的新文件会在原文件名后添加'_with_birthday'")
    print("=" * 50)

    while True:
        process_excel_file()

        # 询问是否继续处理其他文件
        continue_choice = input("\n是否继续处理其他文件？(y/n): ").strip().lower()
        if continue_choice != 'y':
            print("程序结束，再见！")
            break


if __name__ == "__main__":
    main()